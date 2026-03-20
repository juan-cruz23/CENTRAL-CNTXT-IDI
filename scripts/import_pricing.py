"""
Import pricing data from the Excel file 05. Services_05.25.xlsm
into the ServiceTemplate, ProjectCategory, ProjectPhase, and ServiceActivity models.

Usage:
    cd /home/miguelrodriguez/repos/Contexto
    DJANGO_SETTINGS_MODULE=config.settings.local ./venv/bin/python scripts/import_pricing.py
"""

import os
import sys
from collections import defaultdict
from decimal import Decimal

import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings.local")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
django.setup()

import openpyxl

from apps.accounts.models import Role
from apps.organizations.models import OperativeLine, BusinessUnit
from apps.services.models import ProjectCategory, ProjectPhase, ServiceActivity, ServiceTemplate

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "Documentacion", "05. Services_05.25.xlsm",
)

# ──────────────────────────────────────────────────────────────────────────────
# Role mapping
# ──────────────────────────────────────────────────────────────────────────────
ROLE_MAP = {
    "VM": "DM",     # Visual Manager → Director de Medios
    "VL": "LA",     # Visual Leader → Líder de Área
    "VS": "VS",     # Visual Support → Visor de Seguimiento
    "DM": "DM",     # Design Manager
    "LA": "LA",     # Leader Architect
    "SUPP": "INT",  # Support Architect → Interventor
    "JUN": "VL",    # Junior Architect → Visor Libre
    "INT": "VS",    # Intern Architect → Visor Seguimiento
}


def get_or_create_role(role_prefix):
    """Map Excel role prefix to DB Role."""
    if not role_prefix:
        return None
    role_prefix = str(role_prefix)
    prefix = role_prefix.split("|")[0].strip().upper()
    code = ROLE_MAP.get(prefix)
    if not code:
        return None
    role, _ = Role.objects.get_or_create(code=code, defaults={"name": code})
    return role


def parse_phase(phase_str):
    """Parse phase string like 'Fase 1' → ProjectPhase."""
    if not phase_str or not isinstance(phase_str, str):
        return None
    phase_str = phase_str.strip()
    for num in (1, 2, 3):
        if str(num) in phase_str:
            phase, _ = ProjectPhase.objects.get_or_create(
                number=num, defaults={"name": f"Fase {num}"}
            )
            return phase
    return None


def ensure_operative_lines():
    """Ensure V.sual (VIS) and D.sign (DV) operative lines exist."""
    bu, _ = BusinessUnit.objects.get_or_create(
        code="MADE", defaults={"name": "MADE - Estudio de Diseño"}
    )
    vis, _ = OperativeLine.objects.get_or_create(
        code="VIS", defaults={"name": "Visualización", "business_unit": bu}
    )
    dv, _ = OperativeLine.objects.get_or_create(
        code="DV", defaults={"name": "Diseño y Visual", "business_unit": bu}
    )
    return vis, dv


def import_services_catalog(wb):
    """Import from the 'Servicios' sheet — master list of all services."""
    vis_line, dv_line = ensure_operative_lines()
    ws = wb["Servicios"]

    created = 0
    updated = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        code = row[5]
        if not code or not isinstance(code, str) or code == "#N/A":
            continue

        line_code = row[2]  # 'V' or 'D'
        cat_name = row[7]   # e.g. 'V0. Modelación 3D'
        phase_str = row[8]  # e.g. 'Fase 1'
        name = row[9]       # service name
        description = row[10] or ""

        if not name:
            continue

        # Determine operative line
        if line_code == "V":
            operative_line = vis_line
        elif line_code == "D":
            operative_line = dv_line
        else:
            operative_line = None

        # Category
        if cat_name:
            cat_code = cat_name.split(".")[0].strip() if "." in cat_name else cat_name[:5]
            category, _ = ProjectCategory.objects.get_or_create(
                code=cat_code, defaults={"name": cat_name}
            )
            # Update name if it changed
            if category.name != cat_name:
                category.name = cat_name
                category.save()
        else:
            category, _ = ProjectCategory.objects.get_or_create(
                code="GEN", defaults={"name": "General"}
            )

        # Phase
        phase = parse_phase(phase_str)
        if not phase:
            phase, _ = ProjectPhase.objects.get_or_create(
                number=1, defaults={"name": "Fase 1"}
            )

        # Create or update ServiceTemplate
        st, was_created = ServiceTemplate.objects.update_or_create(
            code=code,
            defaults={
                "name": name.strip(),
                "category": category,
                "phase": phase,
                "description": description.strip()[:2000] if description else "",
                "operative_line": operative_line,
                "is_active": True,
            },
        )

        if was_created:
            created += 1
        else:
            updated += 1

    print(f"[Servicios] Created: {created}, Updated: {updated}")
    return created + updated


def import_hours_and_costs(wb):
    """Import hours and costs from V.sual DATA and D.sign DATA sheets."""
    # ── V.sual DATA ──
    ws_v = wb["V.sual DATA"]
    v_hours = defaultdict(float)
    v_costs = defaultdict(float)
    for row in ws_v.iter_rows(min_row=4, values_only=True):
        code = row[1]
        h = row[12]      # TIEMPO (HORAS)
        cost = row[31]   # VALOR TOTAL SERVICIO
        if code and isinstance(code, str) and code != "#N/A":
            if h and isinstance(h, (int, float)):
                v_hours[code] += float(h)
            if cost and isinstance(cost, (int, float)):
                v_costs[code] += float(cost)

    updated_v = 0
    for code, hours in v_hours.items():
        try:
            st = ServiceTemplate.objects.get(code=code)
            st.estimated_hours = Decimal(str(round(hours, 2)))
            st.base_unit_price = Decimal(str(round(v_costs.get(code, 0), 2)))
            st.save()
            updated_v += 1
        except ServiceTemplate.DoesNotExist:
            print(f"  [WARN] V.sual service {code} not in catalog")

    print(f"[V.sual DATA] Updated hours/costs for {updated_v} services")

    # ── D.sign DATA ──
    ws_d = wb["D.sign DATA"]
    d_hours = defaultdict(float)
    d_costs = defaultdict(float)
    for row in ws_d.iter_rows(min_row=5, values_only=True):
        code = row[2]     # col C = CÓDIGO
        h = row[13]       # col N = TIEMPO (HORAS)
        cost = row[32]    # col AG = VALOR TOTAL SERVICIO
        if code and isinstance(code, str) and code not in ("#N/A", "TOTAL"):
            if h and isinstance(h, (int, float)):
                d_hours[code] += float(h)
            if cost and isinstance(cost, (int, float)):
                d_costs[code] += float(cost)

    updated_d = 0
    for code, hours in d_hours.items():
        try:
            st = ServiceTemplate.objects.get(code=code)
            st.estimated_hours = Decimal(str(round(hours, 2)))
            st.base_unit_price = Decimal(str(round(d_costs.get(code, 0), 2)))
            st.save()
            updated_d += 1
        except ServiceTemplate.DoesNotExist:
            print(f"  [WARN] D.sign service {code} not in catalog")

    print(f"[D.sign DATA] Updated hours/costs for {updated_d} services")


def import_activities(wb):
    """Import service activities (acciones) from V.sual DATA and D.sign DATA."""
    total_created = 0

    # ── V.sual DATA activities ──
    ws_v = wb["V.sual DATA"]
    current_code = None
    order_counter = defaultdict(int)

    for row in ws_v.iter_rows(min_row=4, values_only=True):
        code = row[1]
        action = row[10]    # ACCIONES CLAVES
        role_str = row[11]  # RESPONSABLE
        hours = row[12]     # TIEMPO (HORAS)

        if code and isinstance(code, str) and code != "#N/A":
            current_code = code

        if not current_code or not action:
            continue

        try:
            st = ServiceTemplate.objects.get(code=current_code)
        except ServiceTemplate.DoesNotExist:
            continue

        order_counter[current_code] += 1
        role = get_or_create_role(role_str) if role_str else None
        est_hours = Decimal(str(round(float(hours), 2))) if hours and isinstance(hours, (int, float)) else Decimal("0")

        _, created = ServiceActivity.objects.get_or_create(
            service_template=st,
            order=order_counter[current_code],
            defaults={
                "name": str(action).strip()[:300],
                "responsible_role": role,
                "estimated_hours": est_hours,
            },
        )
        if created:
            total_created += 1

    print(f"[V.sual Activities] Created: {total_created}")

    # ── D.sign DATA activities ──
    ws_d = wb["D.sign DATA"]
    current_code = None
    order_counter_d = defaultdict(int)
    created_d = 0

    for row in ws_d.iter_rows(min_row=5, values_only=True):
        code = row[2]       # CÓDIGO
        action = row[11]    # ACCIONES CLAVES
        role_str = row[12]  # RESPONSABLE
        hours = row[13]     # TIEMPO (HORAS)

        if code and isinstance(code, str) and code not in ("#N/A", "TOTAL"):
            current_code = code

        if not current_code or not action:
            continue

        try:
            st = ServiceTemplate.objects.get(code=current_code)
        except ServiceTemplate.DoesNotExist:
            continue

        order_counter_d[current_code] += 1
        role = get_or_create_role(role_str) if role_str else None
        est_hours = Decimal(str(round(float(hours), 2))) if hours and isinstance(hours, (int, float)) else Decimal("0")

        _, created = ServiceActivity.objects.get_or_create(
            service_template=st,
            order=order_counter_d[current_code],
            defaults={
                "name": str(action).strip()[:300],
                "responsible_role": role,
                "estimated_hours": est_hours,
            },
        )
        if created:
            created_d += 1

    print(f"[D.sign Activities] Created: {created_d}")
    return total_created + created_d


def print_summary():
    """Print summary of imported data."""
    total = ServiceTemplate.objects.filter(is_active=True).count()
    vis = ServiceTemplate.objects.filter(operative_line__code="VIS", is_active=True).count()
    dv = ServiceTemplate.objects.filter(operative_line__code="DV", is_active=True).count()
    cats = ProjectCategory.objects.count()
    activities = ServiceActivity.objects.count()

    from django.db.models import Sum
    total_hours = ServiceTemplate.objects.filter(is_active=True).aggregate(
        h=Sum("estimated_hours")
    )["h"] or 0
    total_value = ServiceTemplate.objects.filter(is_active=True).aggregate(
        v=Sum("base_unit_price")
    )["v"] or 0

    print("\n" + "=" * 60)
    print("RESUMEN DE IMPORTACIÓN")
    print("=" * 60)
    print(f"Servicios totales:    {total}")
    print(f"  V.sual:             {vis}")
    print(f"  D.sign:             {dv}")
    print(f"Categorías:           {cats}")
    print(f"Actividades:          {activities}")
    print(f"Horas totales:        {total_hours:,.0f}h")
    print(f"Valor total:          ${total_value:,.0f}")
    print("=" * 60)


def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    print("\n1/3 Importing service catalog...")
    import_services_catalog(wb)

    print("\n2/3 Importing hours and costs...")
    import_hours_and_costs(wb)

    print("\n3/3 Importing activities...")
    import_activities(wb)

    print_summary()
    print("\nDone!")


if __name__ == "__main__":
    main()
