"""
Export / import de ServiceTemplate con árbol completo de entregables → actividades → acciones.
Formato: una fila por acción, columnas de servicio repetidas.
"""
from collections import OrderedDict
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Definición de columnas (orden del Excel) ─────────────────────────────────
COLUMNS = [
    "FASES",
    "CÓDIGO",
    "CATEGORÍA",
    "SUB CATEGORÍA",
    "NOMBRE DE SERVICIO",
    "ENTREGABLE",
    "UND ENTREGABLE",
    "CANTIDAD",
    "ACTIVIDADES CLAVES",
    "# ACCIONES",
    "ACCIONES CLAVES",
    "RESPONSABLE",
    "TIEMPO (HORAS)",
    "HONORARIOS POR HORA ($COP)",
    "SUBTOTAL HONORARIOS ($COP)",
    "HARDWARE ($ DEPRECIACIÓN)",
    "SOFTWARE (LICENCIAS)",
    "CONSUMIBLES ($ INSUMOS)",
    "SUBCONTRATOS",
    "COSTO DIRECTO ($COP)",
    "PRORRATEO GASTOS ($COP)",
    "COSTOS OPERACIONALES ($COP)",
    "DESFASE (%)",
    "DESFASE ($COP)",
    "UTILIDAD (%)",
    "UTILIDAD ($COP)",
    "VALOR NETO ($COP)",
    "MARGEN DE NEGOCIACIÓN (%)",
    "NEGOCIACIÓN ($COP)",
    "ICA ($COP)",
    "4 * 1000 ($COP)",
    "VALOR TOTAL SERVICIO ($COP)",
]

# Columnas que son input editable en el modelo (las demás son calculadas)
INPUT_COLS = {
    "HONORARIOS POR HORA ($COP)",
    "HARDWARE ($ DEPRECIACIÓN)",     # total → se divide entre horas al importar
    "SOFTWARE (LICENCIAS)",
    "CONSUMIBLES ($ INSUMOS)",
    "SUBCONTRATOS",
    "DESFASE (%)",
    "UTILIDAD (%)",
    "MARGEN DE NEGOCIACIÓN (%)",
}

# Columnas calculadas (sombreado más tenue en el export)
CALC_COLS = {c for c in COLUMNS if c not in INPUT_COLS and c not in {
    "FASES", "CÓDIGO", "CATEGORÍA", "SUB CATEGORÍA", "NOMBRE DE SERVICIO",
    "ENTREGABLE", "UND ENTREGABLE", "CANTIDAD", "ACTIVIDADES CLAVES",
    "# ACCIONES", "ACCIONES CLAVES", "RESPONSABLE", "TIEMPO (HORAS)",
}}

# Ancho de cada columna (mismo orden que COLUMNS)
COL_WIDTHS = [
    12, 12, 22, 18, 38,   # FASES … NOMBRE
    32, 12, 8,             # ENTREGABLE … CANTIDAD
    32, 8, 48, 20, 12,    # KACT … HORAS
    18, 18,                # HONORARIOS HORA / SUBTOTAL
    16, 16, 16, 14,        # HW / SW / CONS / SUBCONTR
    16, 16, 20,            # CD / PRORRAT / COSTOS OP
    10, 14, 10, 14, 16,    # DESFASE % $ / UTIL % $ / NETO
    16, 14, 12, 12, 18,    # NEGOC % $ / ICA / 4X1000 / TOTAL
]


# ── EXPORT ───────────────────────────────────────────────────────────────────

def export_services(queryset=None):
    """Genera un Workbook openpyxl con todos los servicios del queryset."""
    from apps.services.models import ServiceTemplate
    from domain.financials.cost_allocation import get_current_prorrateo_rate

    if queryset is None:
        queryset = ServiceTemplate.objects.all()

    qs = (
        queryset
        .select_related("phase", "category", "subcategory", "responsible_role", "operative_line")
        .prefetch_related(
            "deliverables__key_activities__actions__responsible_role",
            "hardwares",
            "softwares",
        )
        .order_by("code")
    )

    # Materializa una vez para que cached_property persista a lo largo del
    # loop (si iteráramos `qs` dos veces, cada pasada daría instancias nuevas).
    services = list(qs)

    # Pre-cachea la tarifa de prorrateo por línea operativa única: evita 3
    # queries × N plantillas dentro del motor de cost_allocation.
    rate_cache: dict = {}
    for st in services:
        line_code = st.operative_line.code if st.operative_line_id else None
        if line_code not in rate_cache:
            rate_cache[line_code] = get_current_prorrateo_rate(line_code)
        st.__dict__["effective_prorrateo_rate"] = rate_cache[line_code]

    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios"

    # ── Cabecera ──
    _write_header(ws)

    row_num = 2
    for st in services:
        deliverables = sorted(st.deliverables.all(), key=lambda d: (d.order, d.name))
        if not deliverables:
            _write_row(ws, row_num, st, None, None, None, None)
            row_num += 1
            continue

        for deliv in deliverables:
            kacts = sorted(deliv.key_activities.all(), key=lambda k: (k.order, k.name))
            if not kacts:
                _write_row(ws, row_num, st, deliv, None, None, None)
                row_num += 1
                continue

            for kact in kacts:
                actions = sorted(kact.actions.all(), key=lambda a: a.order)
                if not actions:
                    _write_row(ws, row_num, st, deliv, kact, None, None)
                    row_num += 1
                    continue

                for act_num, act in enumerate(actions, 1):
                    _write_row(ws, row_num, st, deliv, kact, act, act_num)
                    row_num += 1

    # Freeze y filtro automático
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    return wb


def _write_header(ws):
    h_font  = Font(bold=True, color="FFFFFF", size=9)
    h_fill  = PatternFill("solid", fgColor="1F3864")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = h_font
        cell.fill   = h_fill
        cell.alignment = h_align

    ws.row_dimensions[1].height = 42

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_row(ws, row_num, st, deliv, kact, act, act_num):
    def d(v):
        try:
            return float(v) if v else 0
        except (TypeError, ValueError):
            return 0

    values = [
        st.phase.name if st.phase else "",
        st.code,
        st.category.name if st.category else "",
        st.subcategory.name if st.subcategory else "",
        st.name,
        deliv.name if deliv else "",
        deliv.unit if deliv else "",
        float(deliv.quantity) if deliv else "",
        kact.name if kact else "",
        act_num if act_num else "",
        act.name if act else "",
        act.responsible_role.code if (act and act.responsible_role) else "",
        float(act.estimated_hours) if act else "",
        d(st.hourly_rate),
        d(st.subtotal_honorarios),
        d(st.hardware_total),
        d(st.software_total),
        d(st.consumables_total),
        d(st.subcontracts),
        d(st.costo_directo),
        d(st.prorrateo_gastos),
        d(st.costos_operacionales),
        d(st.contingency_pct),
        d(st.desfase_value),
        d(st.utility_pct),
        d(st.utility_value),
        d(st.valor_neto),
        d(st.negotiation_pct),
        d(st.negotiation_value),
        d(st.ica),
        d(st.gmf_4x1000),
        d(st.valor_total_servicio),
    ]

    # Relleno alternado para legibilidad
    even_fill = PatternFill("solid", fgColor="EEF2FF")
    # Relleno para columnas calculadas (lectura)
    calc_fill = PatternFill("solid", fgColor="F0F9FF")

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        col_name = COLUMNS[col_idx - 1]
        if col_name in CALC_COLS:
            cell.fill = calc_fill
            cell.font = Font(color="555555", size=9)
        elif row_num % 2 == 0:
            cell.fill = even_fill
        # Formato numérico para columnas de pesos
        if col_name.endswith("($COP)") or col_name == "SUBCONTRATOS":
            cell.number_format = '#,##0.00'
        elif col_name in {"DESFASE (%)", "UTILIDAD (%)", "MARGEN DE NEGOCIACIÓN (%)"}:
            cell.number_format = '0.00'
        elif col_name in {"CANTIDAD", "TIEMPO (HORAS)", "# ACCIONES"}:
            cell.number_format = '0.00'


# ── IMPORT ───────────────────────────────────────────────────────────────────

def import_services(file_obj, inactivate_missing=True):
    """
    Parsea un Excel con la estructura de export_services().
    Estrategia:
      - Upsert por CÓDIGO: si existe → actualiza y reconstruye el árbol
      - Si no existe → crea
      - Si inactivate_missing=True: servicios no en el archivo → is_active=False

    Retorna dict: {created, updated, inactivated, errors: [str]}
    """
    from apps.services.models import (
        Deliverable, KeyActivity, ProjectCategory, ProjectPhase,
        ServiceActivity, ServiceSubCategory, ServiceTemplate,
    )
    from apps.accounts.models import Role

    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception as exc:
        return {"created": 0, "updated": 0, "inactivated": 0, "errors": [f"No se pudo leer el archivo: {exc}"]}

    ws = wb.active

    # ── Mapear columna → índice 0-based ──
    raw_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = {}
    for idx, h in enumerate(raw_headers):
        if h is not None:
            col_map[str(h).strip()] = idx

    missing = [c for c in COLUMNS if c not in col_map]
    if missing:
        return {
            "created": 0, "updated": 0, "inactivated": 0,
            "errors": [f"Columnas faltantes en el archivo: {', '.join(missing)}"]
        }

    def get(row_tuple, col_name):
        idx = col_map.get(col_name)
        if idx is None:
            return ""
        val = row_tuple[idx]
        return str(val).strip() if val is not None else ""

    def get_dec(row_tuple, col_name):
        raw = get(row_tuple, col_name)
        try:
            return Decimal(str(raw).replace(",", ".")) if raw else Decimal("0")
        except (InvalidOperation, ValueError):
            return Decimal("0")

    # ── Leer todas las filas de datos ──
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v for v in row if v not in (None, "")):
            continue
        all_rows.append(row)

    # ── Agrupar por CÓDIGO (mantener orden de aparición) ──
    services_data: dict[str, list] = OrderedDict()
    rows_without_code = 0
    for row in all_rows:
        code = get(row, "CÓDIGO")
        if not code:
            rows_without_code += 1
            continue
        services_data.setdefault(code, []).append(row)

    # ── Caches de maestros ──
    phases       = {p.name.strip(): p for p in ProjectPhase.objects.all()}
    categories   = {c.name.strip(): c for c in ProjectCategory.objects.all()}
    subcategories = {s.name.strip(): s for s in ServiceSubCategory.objects.all()}
    roles_by_code = {r.code.strip(): r for r in Role.objects.filter(is_active=True) if r.code}
    roles_by_name = {r.name.strip(): r for r in Role.objects.filter(is_active=True)}

    def find_role(code_or_name):
        if not code_or_name:
            return None
        return roles_by_code.get(code_or_name) or roles_by_name.get(code_or_name)

    created = updated = 0
    errors  = []
    if rows_without_code:
        errors.append(f"{rows_without_code} fila(s) omitidas por CÓDIGO vacío — completa la columna CÓDIGO para crear nuevos servicios.")
    processed_codes = set()

    for code, rows in services_data.items():
        processed_codes.add(code)
        first = rows[0]

        phase_name  = get(first, "FASES")
        cat_name    = get(first, "CATEGORÍA")
        subcat_name = get(first, "SUB CATEGORÍA")
        st_name     = get(first, "NOMBRE DE SERVICIO")

        if not cat_name:
            errors.append(f"[{code}] Sin CATEGORÍA → omitido.")
            continue

        category = categories.get(cat_name)
        if not category:
            errors.append(f"[{code}] Categoría '{cat_name}' no existe → omitido.")
            continue

        phase      = phases.get(phase_name) if phase_name else None
        subcategory = subcategories.get(subcat_name) if subcat_name else None

        # Sumar horas totales de acciones
        total_hours = sum(get_dec(r, "TIEMPO (HORAS)") for r in rows if get(r, "ACCIONES CLAVES"))
        if not total_hours:
            total_hours = Decimal("0")

        # Tasas por hora (el Excel exporta totales → divide por horas)
        hourly_rate = get_dec(first, "HONORARIOS POR HORA ($COP)")

        def rate_from_total(col):
            total = get_dec(first, col)
            if total_hours:
                return (total / total_hours).quantize(Decimal("0.01"))
            return Decimal("0")

        hw_per_h   = rate_from_total("HARDWARE ($ DEPRECIACIÓN)")
        sw_per_h   = rate_from_total("SOFTWARE (LICENCIAS)")
        cons_per_h = rate_from_total("CONSUMIBLES ($ INSUMOS)")
        subcontr   = get_dec(first, "SUBCONTRATOS")
        cont_pct   = get_dec(first, "DESFASE (%)")   or Decimal("15")
        util_pct   = get_dec(first, "UTILIDAD (%)")  or Decimal("20")
        neg_pct    = get_dec(first, "MARGEN DE NEGOCIACIÓN (%)") or Decimal("5")

        st_defaults = dict(
            name=st_name,
            category=category,
            phase=phase,
            subcategory=subcategory,
            estimated_hours=total_hours,
            hourly_rate=hourly_rate,
            hardware_cost_per_hour=hw_per_h,
            software_cost_per_hour=sw_per_h,
            consumables_per_hour=cons_per_h,
            subcontracts=subcontr,
            contingency_pct=cont_pct,
            utility_pct=util_pct,
            negotiation_pct=neg_pct,
            is_active=True,
        )

        # Savepoint por servicio: con ATOMIC_REQUESTS=True un error envenena
        # toda la transacción del request; transaction.atomic() anidado crea
        # un savepoint que se puede rollback sin afectar el resto del import.
        try:
            with transaction.atomic():
                st, is_new = ServiceTemplate.objects.get_or_create(code=code, defaults=st_defaults)
                if is_new:
                    created += 1
                else:
                    for k, v in st_defaults.items():
                        setattr(st, k, v)
                    st.save()
                    updated += 1
                    # Reconstruir árbol completo: actividades primero (key_activity SET_NULL)
                    st.activities.all().delete()
                    st.deliverables.all().delete()
        except Exception as exc:
            errors.append(f"[{code}] Error al guardar servicio: {exc}")
            continue

        # ── Reconstruir árbol: Entregables → KActs → Acciones ──
        # Agrupar por (nombre entregable, unidad, cantidad) preservando orden
        deliv_groups: dict[tuple, list] = OrderedDict()
        for row in rows:
            d_name = get(row, "ENTREGABLE")
            if not d_name:
                continue
            d_unit = get(row, "UND ENTREGABLE")
            d_qty  = str(get_dec(row, "CANTIDAD") or "1")
            key    = (d_name, d_unit, d_qty)
            deliv_groups.setdefault(key, []).append(row)

        for d_order, ((d_name, d_unit, d_qty_str), d_rows) in enumerate(deliv_groups.items(), 1):
            try:
                with transaction.atomic():
                    deliv = Deliverable.objects.create(
                        service_template=st,
                        name=d_name,
                        unit=d_unit,
                        quantity=Decimal(d_qty_str),
                        order=d_order,
                    )
            except Exception as exc:
                errors.append(f"[{code}] Error entregable '{d_name}': {exc}")
                continue

            # Agrupar por actividad clave
            kact_groups: dict[str, list] = OrderedDict()
            for row in d_rows:
                ka_name = get(row, "ACTIVIDADES CLAVES")
                if not ka_name:
                    continue
                kact_groups.setdefault(ka_name, []).append(row)

            for ka_order, (ka_name, ka_rows) in enumerate(kact_groups.items(), 1):
                try:
                    with transaction.atomic():
                        kact = KeyActivity.objects.create(
                            deliverable=deliv, name=ka_name, order=ka_order
                        )
                except Exception as exc:
                    errors.append(f"[{code}/{d_name}] Error actividad '{ka_name}': {exc}")
                    continue

                for act_order, row in enumerate(ka_rows, 1):
                    act_name = get(row, "ACCIONES CLAVES")
                    if not act_name:
                        continue
                    role_code = get(row, "RESPONSABLE")
                    role      = find_role(role_code)
                    hours     = get_dec(row, "TIEMPO (HORAS)")

                    try:
                        with transaction.atomic():
                            ServiceActivity.objects.create(
                                service_template=st,
                                key_activity=kact,
                                name=act_name,
                                order=act_order,
                                responsible_role=role,
                                estimated_hours=hours,
                            )
                    except Exception as exc:
                        errors.append(f"[{code}/{ka_name}] Error acción '{act_name}': {exc}")

    # ── Inactivar servicios no presentes en el archivo ──
    inactivated = 0
    if inactivate_missing and processed_codes:
        qs_inactive = ServiceTemplate.objects.filter(is_active=True).exclude(code__in=processed_codes)
        inactivated = qs_inactive.count()
        qs_inactive.update(is_active=False)

    return {
        "created": created,
        "updated": updated,
        "inactivated": inactivated,
        "errors": errors,
    }
